from openpyxl import Workbook, load_workbook
from fastapi import UploadFile
from io import BytesIO
from datetime import date
from app.database import models
from typing import List
from sqlalchemy.orm import Session


def build_orders_excel(orders: List[models.Order]) -> BytesIO:
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    headers = ["Order Number", "Client", "Beverage Name", "MRP", "Quantity", "Cases", "Discount", "Total Price", "Status"]
    ws.append(headers)

    for order in orders:
        for item in order.items:
            ws.append([
                order.order_name,
                order.client.outlet_name,
                item.beverage.product_brand,
                item.mrp,
                item.quantity,
                item.cases,
                item.discount,
                item.total_price,
                "Delivered" if item.order.status == 1 else "Pending"
            ])
            
    # auto-width columns
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
        ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def decode_orders_excel(file: UploadFile, db: Session):
    wb = load_workbook(filename=BytesIO(file.file.read()), data_only=True)
    ws = wb.active
    
    last_order = None
    delivery_status = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        order_name = str(row[0]).strip()

        order = db.query(models.Order).filter(models.Order.order_name == order_name).first()

        if not order:
            continue

        last_order = order
        if len(row) > 8 and str(row[8]).strip().lower() == "delivered":
            delivery_status += 1

    if last_order:
        total_items = len(last_order.items)
        undelivered_items = total_items - delivery_status
        if undelivered_items <= 0:
            last_order.status = 1
            db.commit()
    else:
        "Orders Partially delivered"

    return "Orders updated successfully"
